from pprint import pprint
import os
from yattag import Doc
from yattag import indent
import openpyxl
from collections import defaultdict

#opening excel data for extraction
wb = openpyxl.load_workbook('database.xlsx')
#selecting sheet name containing data
sheet = wb.get_sheet_by_name('data')
#creating dictionary of row data
row_data_dict = defaultdict(lambda : defaultdict(dict))

#iterating through the data rows
for i_row in range(1,sheet.max_row):
    row_data = list(map(lambda x: x.value,sheet.rows[i_row]))
    for i in range(0,(sheet.max_column-2),2):
        row_data_dict[row_data[0]][row_data[1]][row_data[2+i]] = row_data[3+i]
pprint(row_data_dict)

#List of product
itemList = row_data_dict.keys()

#items -  Variable which stores product names in sorted order.
items = (sorted(itemList))

#pretty prints the list of items.
pprint(items);


#for loop which iterates through list of items
for item in items:
    #filename - removes spaces from item's name.
    filename = str(item.replace(' ',''))

    #using yattag's ability to make html docs
    doc, tag, text = Doc().tagtext()

    
    with tag('html'):
        with tag('head'):
            doc.stag("meta", charset="utf-8")
            doc.stag("meta", name="viewport", content="width=device-width, initial-scale=1")
            with tag('title'):
                text('TPCPL - '+item)
            doc.stag("link", rel="stylesheet", type="text/css",href="../../CSS/hover.css")    
            doc.stag("link", rel="stylesheet", type="text/css",href="../../CSS/custom.css")
            doc.stag("link", rel="stylesheet", type="text/css",href="../../CSS/jquery-ui.min.css")
            doc.stag("link", rel="stylesheet", type="text/css",href="../CSS/product.css")
            
        with tag('body'):
            #MAIN HEADING
            with tag("a", id="title",klass="container hvr-sweep-to-left",href="../../index.html"):
                text("PRODUCT DATA MANAGER");

            #MAIN BODY
            with tag("div", klass="heading"):
                doc.stag("img", src=str(filename)+"Logo.png",klass="logo")
                text(item)
            with tag("ul", klass="unordered-list"):
                for model in sorted(row_data_dict[item].keys()):
                    modelConcat = model.replace(" ","");
                    with tag("a"):
                        with tag("li", klass="list hvr-sweep-to-top", id="list-"+modelConcat,onclick="openData(id)"):
                            doc.stag("img",src=str(modelConcat)+".png", klass="list-image")
                            with tag("div", klass="item-name"):
                                text(str(model))
            for model in sorted(row_data_dict[item].keys()):
                modelConcat = model.replace(" ","");
                with tag("div", id="data-list-"+modelConcat, klass="data-display"):
                    with tag("div",id="data-list-"+modelConcat+"-x", onclick="closeData(id)", klass="data-display-x hvr-pulse-shrink"):
                        text("x")
                    with tag("div", klass="data-display-head"):
                        doc.stag("img", src=str(filename)+"Logo.png", klass="logo")
                        text(item)
                    with tag("div", klass="unordered-list", id="data-background"):
                        with tag("div", klass="list hvr-sweep-to-top"):
                            doc.stag("img", src=str(modelConcat)+".png", klass="list-imag")
                            with tag("div", klass="item-name"):
                                text(model)
                        doc.stag("br")
                        with tag("ul", klass="unordered-list"):
                            for unitsRange, rate in sorted(row_data_dict[item][model].items()):
                                with tag("li", klass="data-list hvr-sweep-to-top"):
                                    with tag("div", klass="rate-font"):
                                        text("Rs. "+str(rate))
                                    text(unitsRange)    
                        
            #JS TAGS
            with tag("script", src="../../JS/jquery.js"):
                pass
            with tag("script", src="../../JS/jquery-ui.min.js"):
                pass
            with tag("script", src="../../JS/completelist.js"):
                pass
            with tag("script", src="../../JS/droplist.js"):
                pass
            with tag("script", src="../JS/image.js"):
                pass
           

    #indent the resulting document
    result = indent(
        doc.getvalue(),
        indentation = '    ',
        newline = '\r\n'
    )
           
    try:
        if not os.path.exists(filename):
            os.makedirs(filename)
        with open(os.path.join(filename, filename+'.html'), 'w+') as temp_file:
            temp_file.write(result)
    except Exception as e:
        print(e)


